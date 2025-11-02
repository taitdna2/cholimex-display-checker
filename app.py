# app.py
from __future__ import annotations
import io, os, re, json
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from collections import defaultdict

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st
import zipfile

# =========================
# DEFAULT CONFIG (fallback)
# =========================
DEFAULT_CONFIG = {
    "muc_toi_thieu": {
        "NMCD": 150000,
        "DHLM": 100000,
        "KOS_XXTG": 300000,
        "LTLKC": 80000,
        "GVIG": 300000,
        "GVIG_BMTR": 300000,
        "KOS_XXTG_BS": 200000,
        "CAKOS": 50000,
        "XBM_MN": 36000,
        "XBM_MB": 36000
    },
    "program_names": {
        "NMCD": "Trưng bày Nước mắm Cholimex 30, 35, 40 độ đạm 500ml + 750ml",
        "DHLM": "Trưng bày Dầu hào 820g, Nước tương Lên men 700ml",
        "LTLKC": "Trưng bày Xốt Lẩu thái 280g & Xốt Lẩu kim chi 280g",
        "KOS_XXTG": "Trưng bày cá KOS và Xúc xích - Miền Nam",
        "KOS_XXTG_BS": "Trưng bày cá KOS và Xúc xích - Miền Bắc & Bắc Miền Trung",
        "XBM_MN": "Trưng bày Xe Bánh Mì - Miền Nam",
        "XBM_MB": "Trưng bày Xe Bánh Mì - Miền Bắc",
        "CAKOS": "Trưng bày cá KOS",
        "GVIG": "Trưng bày Gia vị gói - Miền Bắc",
        "GVIG_BMTR": "Trưng bày Gia vị gói - Bắc Miền Trung"
    },
    "region_map": {
        "HCME": ["HCM", "MD"],
        "MTRUNG": ["MTR", "MB_MT3"],
        "MTAY": ["MTA"],
        "MBAC": ["MB"],
        "TOAN_QUOC": "ALL"
    },
    "xbm_map": {"M70": "XBM_MN", "M110": "XBM_MN", "M80": "XBM_MB", "M120": "XBM_MB"}
}

# =========================
# CONFIG LOADER (optional JSON next to app)
# =========================
APP_DIR = os.getcwd()
CONFIG_DIR = os.path.join(APP_DIR, "config")

def _load_json(path: str) -> Optional[dict]:
    try:
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return None

def load_config() -> Dict:
    cfg = DEFAULT_CONFIG.copy()
    mt = _load_json(os.path.join(CONFIG_DIR, "muc_toi_thieu.json"))
    pn = _load_json(os.path.join(CONFIG_DIR, "program_names.json"))
    rm = _load_json(os.path.join(CONFIG_DIR, "region_map.json"))
    xb = _load_json(os.path.join(CONFIG_DIR, "xbm_map.json"))
    if mt: cfg["muc_toi_thieu"] = mt
    if pn: cfg["program_names"] = pn
    if rm: cfg["region_map"] = rm
    if xb: cfg["xbm_map"] = xb
    return cfg

CFG = load_config()
MUC_TOI_THIEU = CFG["muc_toi_thieu"]
PROGRAM_NAMES = CFG["program_names"]
REGION_MAP = CFG["region_map"]
XBM_MAP = CFG["xbm_map"]

# =============== Common utils ===============
def parse_giai_to_dt(giai: str) -> datetime:
    """Dò định dạng 'Giai đoạn' và trả về datetime(YYYY, MM, 1)."""
    s = str(giai).strip()
    # Tháng 11/2025, T11/2025, 11/2025
    m = re.search(r'(\d{1,2})\D+([12]\d{3})', s)
    if m:
        mm, yy = int(m.group(1)), int(m.group(2))
        return datetime(yy, mm, 1)
    # 2025-11, 2025/11
    m = re.search(r'([12]\d{3})\D+(\d{1,2})', s)
    if m:
        yy, mm = int(m.group(1)), int(m.group(2))
        return datetime(yy, mm, 1)
    # T11 (mặc định năm hiện tại)
    m = re.fullmatch(r'[Tt]?(\d{1,2})', s)
    if m:
        mm = int(m.group(1))
        yy = datetime.now().year
        return datetime(yy, mm, 1)
    # fallback: cố gắng parse tự do
    try:
        return pd.to_datetime(s, dayfirst=True).to_pydatetime().replace(day=1)
    except Exception:
        raise ValueError(f"Không nhận diện được Giai đoạn: {s}")

def detect_ct_from_filename(fname: str) -> Optional[str]:
    keys = list(MUC_TOI_THIEU.keys())
    key_pat = "|".join(re.escape(k) for k in keys)
    m = re.search(key_pat, fname, flags=re.I)
    if m:
        return m.group(0).upper()
    return None

def detect_ct_from_content(df: pd.DataFrame) -> Optional[str]:
    if "Mức đăng ký" not in df.columns:
        return None
    vals = df["Mức đăng ký"].dropna().astype(str).str.strip().unique().tolist()
    if not vals:
        return None
    # Nếu là XBM mã Mxx -> map
    mapped = set()
    for v in vals:
        v_up = v.upper()
        if v_up in XBM_MAP:
            mapped.add(XBM_MAP[v_up])
        elif v_up in MUC_TOI_THIEU:
            mapped.add(v_up)
    if len(mapped) == 1:
        return list(mapped)[0]
    # nếu nhiều CT trong một file -> không chắc chắn
    return None

# =============== Excel styling ===============
from openpyxl import Workbook

def style_excel(writer: pd.ExcelWriter, sheet_name: str) -> None:
    ws = writer.sheets[sheet_name]
    thin_border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 30

    # Định dạng tiền cho các cột "Doanh số tích lũy *" & "Ngưỡng tối thiểu"
    money_cols = set()
    for idx, cell in enumerate(ws[1], start=1):
        v = str(cell.value) if cell.value is not None else ""
        if v.startswith("Doanh số tích lũy") or v == "Ngưỡng tối thiểu":
            money_cols.add(idx)

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for c in r:
            c.border = thin_border
            c.alignment = Alignment(vertical="center")
            if c.col_idx in money_cols:
                c.number_format = "#,##0"

    # Tô màu theo 'Kết quả'
    col_kq = None
    for i, c in enumerate(ws[1], start=1):
        if c.value == "Kết quả":
            col_kq = i
            break
    if col_kq:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_kq)
            if cell.value == "Không đạt":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif cell.value == "Không xét":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Auto width
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

def tao_bao_cao_tonghop(writer: pd.ExcelWriter, bao_cao_data: List[List]) -> None:
    ws = writer.book.create_sheet("BaoCao_TongHop")
    ws.merge_cells("A1:F1")
    ws["A1"] = "BÁO CÁO QUẦY TRƯNG BÀY KHÔNG ĐẠT DOANH SỐ TỐI THIỂU"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["STT", "Tên chương trình",
               "DOANH SỐ TỐI THIỂU PHÁT SINH/ SUẤT/ THÁNG (VND)",
               "TỔNG SỐ SUẤT TRƯNG BÀY", "SỐ SUẤT KHÔNG ĐẠT", "TỈ LỆ"]
    ws.append(headers)

    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    for i in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=i)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if i in (3, 4): c.fill = fill_yellow
        if i in (5, 6): c.fill = fill_red
        c.border = thin

    for r_idx, row in enumerate(bao_cao_data, start=3):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = thin
            c.alignment = Alignment(horizontal="left" if c_idx == 2 else "center", vertical="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    for col in ("C", "D", "E", "F"):
        ws.column_dimensions[col].width = 22

def tao_bao_cao_huy(writer: pd.ExcelWriter, bao_cao_huy: List[List]) -> None:
    ws = writer.book.create_sheet("BaoCao_Huy")
    ws.merge_cells("A1:C1")
    ws["A1"] = "BÁO CÁO HỦY QUẦY TRƯNG BÀY KHÔNG ĐẠT  DOANH SỐ TỐI THIỂU 2 THÁNG LIÊN TIẾP"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["STT", "Tên chương trình", "TỔNG SỐ SUẤT HỦY TRƯNG BÀY TRÊN HT DMS"]
    ws.append(headers)

    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    for i in range(1, 4):
        c = ws.cell(row=2, column=i)
        c.font = Font(bold=True)
        c.fill = fill_yellow
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

    for r_idx, row in enumerate(bao_cao_huy, start=3):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = thin
            c.alignment = Alignment(horizontal="left" if c_idx == 2 else "center", vertical="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 28

# =========================
# CORE LOGIC (dataframe)
# =========================
def xu_ly_file(file_like, muc_toi_thieu: Dict[str, float], xbm_map: Dict[str, str]) -> Tuple[pd.DataFrame, str]:
    df = pd.read_excel(file_like, header=1, dtype={"Mã khách hàng": str, "Mã NPP": str})
    cols_in = ["Mức đăng ký", "Miền", "Vùng", "Mã NPP", "Tên NPP",
               "Giai đoạn", "Mã NVBH", "Tên NVBH",
               "Mã khách hàng", "Tên khách hàng", "Thứ bán hàng", "Tuyến",
               "Số suất đăng kí", "Doanh số tích lũy hiện tại"]
    df = df[[c for c in cols_in if c in df.columns]].copy()

    df.rename(columns={
        "Mức đăng ký": "MucDK", "Miền": "Mien", "Vùng": "Vung",
        "Mã NPP": "MaNPP", "Tên NPP": "TenNPP",
        "Giai đoạn": "GiaiDoan", "Mã NVBH": "MaNVBH", "Tên NVBH": "TenNVBH",
        "Mã khách hàng": "MaKH", "Tên khách hàng": "TenKH",
        "Thứ bán hàng": "ThuBanHang", "Tuyến": "Tuyen",
        "Số suất đăng kí": "SoSuat", "Doanh số tích lũy hiện tại": "DoanhSo",
    }, inplace=True)

    if "Tuyen" not in df.columns: df["Tuyen"] = None
    if "ThuBanHang" not in df.columns: df["ThuBanHang"] = None

    # Nguỡng = (map theo MucDK hoặc XBM map) * SoSuat
    muc_map = df["MucDK"].astype(str).str.strip().map(xbm_map).fillna(df["MucDK"].astype(str).str.strip())
    base = muc_map.map(muc_toi_thieu).fillna(0).astype(float)
    df["NguongToiThieu"] = base * pd.to_numeric(df["SoSuat"], errors="coerce").fillna(0).astype(float)

    giai = str(df["GiaiDoan"].iloc[0]).strip()
    df[f"SoSuat_{giai}"] = df["SoSuat"]
    df[f"DoanhSo_{giai}"] = df["DoanhSo"]
    df[f"Nguong_{giai}"] = df["NguongToiThieu"]
    return df, giai

def xu_ly_chuong_trinh(
    file_truoc, file_sau,
    muc_toi_thieu: Dict[str, float],
    program_names: Dict[str, str],
    xbm_map: Dict[str, str],
    file_t0: Optional[io.BytesIO] = None,
    filter_ketqua: Optional[set] = None,
    filter_tuyen_tokens: Optional[List[str]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:

    df_t1, g1 = xu_ly_file(file_truoc, muc_toi_thieu, xbm_map)
    df_t2, g2 = xu_ly_file(file_sau, muc_toi_thieu, xbm_map)

    new_in_T1_keys = set()
    g0 = None
    if file_t0:
        df_t0, g0 = xu_ly_file(file_t0, muc_toi_thieu, xbm_map)
        keys_t0 = set(zip(df_t0["MaKH"], df_t0["MucDK"]))
        keys_t1 = set(zip(df_t1["MaKH"], df_t1["MucDK"]))
        new_in_T1_keys = keys_t1 - keys_t0

    df = pd.merge(df_t1, df_t2, on=["MaKH", "MucDK"], how="outer", suffixes=("_T1", "_T2"))
    if file_t0:
        df = df.merge(df_t0[["MaKH", f"SoSuat_{g0}", f"DoanhSo_{g0}"]], on="MaKH", how="left")

    for col in [f"SoSuat_{g1}", f"SoSuat_{g2}", f"DoanhSo_{g1}", f"DoanhSo_{g2}", f"Nguong_{g1}", f"Nguong_{g2}"]:
        if col in df.columns:
            df[col] = df[col].fillna(0)

    def xet(row):
        ds1, ds2 = row.get(f"DoanhSo_{g1}", 0) or 0, row.get(f"DoanhSo_{g2}", 0) or 0
        ss1, ss2 = row.get(f"SoSuat_{g1}", 0) or 0, row.get(f"SoSuat_{g2}", 0) or 0
        n1, n2 = row.get(f"Nguong_{g1}", 0) or 0, row.get(f"Nguong_{g2}", 0) or 0
        key = (row.get("MaKH"), row.get("MucDK"))

        if ss1 > 0 and ss2 == 0: return "XOA", "Tháng trước có tham gia, tháng sau không tham gia"
        if ss1 > 0 and key in new_in_T1_keys: return "Đạt", "Khách mới tháng trước (DS xét chu kỳ 11/T0→10/T1)"
        if ss1 == 0 and ss2 > 0: return "Không xét", "Khách hàng mới tháng sau (không xét kết quả kỳ này)"
        if ss2 > ss1 > 0: return "Đạt", f"Nâng suất từ {int(ss1)} → {int(ss2)} (auto đạt)"
        if ss2 < ss1:
            if (ds1 >= n1) or (ds2 >= n2):
                return "Đạt", f"Giảm suất từ {int(ss1)} → {int(ss2)} (1 trong 2 tháng đủ ngưỡng)"
            else:
                return "Không đạt", f"Giảm suất từ {int(ss1)} → {int(ss2)} (2 tháng đều không đủ ngưỡng)"
        if (ds1 >= n1) or (ds2 >= n2): return "Đạt", ""
        return "Không đạt", "Doanh số 2 tháng liên tiếp không đủ theo yêu cầu"

    df[["KetQua", "GhiChu"]] = df.apply(lambda r: pd.Series(xet(r)), axis=1)

    df_removed = df[df["KetQua"] == "XOA"].copy()
    df_final = df[df["KetQua"] != "XOA"].copy()

    counts = df_final["MaKH"].value_counts()
    multi = counts[counts > 1].index
    df_multi = df_final[df_final["MaKH"].isin(multi)].copy()

    df_final = df_final.sort_values(by=["MaKH", "MaNPP_T2"], na_position="first").drop_duplicates(subset=["MaKH"], keep="last")
    df_removed_multi = df_multi.merge(df_final[["MaKH"]], on="MaKH", how="left", indicator=True)
    df_removed_multi = df_removed_multi[df_removed_multi["_merge"] == "left_only"].drop(columns="_merge")
    df_removed_multi["GhiChu"] = df_removed_multi["MaKH"].map(lambda x: f"Khách hàng xuất hiện {counts[x]} lần, đã giữ bản ghi NPP mới nhất")
    df_removed = pd.concat([df_removed, df_removed_multi], ignore_index=True)

    # Lọc theo Kết quả
    if filter_ketqua is not None and "KetQua" in df_final.columns:
        df_final = df_final[df_final["KetQua"].isin(filter_ketqua)]

    # Lọc theo 'Thứ bán hàng' (fallback 'Tuyến'), không xuất cột 'Tuyến'
    route_col = "ThuBanHang_T2" if "ThuBanHang_T2" in df_final.columns else ("Tuyen_T2" if "Tuyen_T2" in df_final.columns else None)
    if filter_tuyen_tokens and route_col:
        toks = [t.strip().lower() for t in filter_tuyen_tokens if t]
        mask = df_final[route_col].astype(str).str.lower().apply(lambda s: any(tok in s for tok in toks))
        df_final = df_final[mask]

    # Cột xuất ra (BỎ 'Tuyen_T2')
    cols_output = [
        "MucDK", "Mien_T2", "Vung_T2", "MaNPP_T2", "TenNPP_T2",
        "MaNVBH_T2", "TenNVBH_T2", "MaKH", "TenKH_T2", "ThuBanHang_T2",
    ]
    if file_t0:
        cols_output += [f"SoSuat_{g0}", f"DoanhSo_{g0}"]
    cols_output += [
        f"SoSuat_{g1}", f"SoSuat_{g2}",
        f"DoanhSo_{g1}", f"DoanhSo_{g2}",
        f"Nguong_{g2}", "KetQua", "GhiChu",
    ]

    rename_cols = {
        "MucDK": "Mức đăng ký",
        "Mien_T2": "Miền", "Vung_T2": "Vùng",
        "MaNPP_T2": "Mã NPP", "TenNPP_T2": "Tên NPP",
        "MaNVBH_T2": "Mã NVBH", "TenNVBH_T2": "Tên NVBH",
        "MaKH": "Mã khách hàng", "TenKH_T2": "Tên khách hàng",
        "ThuBanHang_T2": "Thứ bán hàng",
        f"SoSuat_{g1}": f"Số suất đăng ký {g1}",
        f"SoSuat_{g2}": f"Số suất đăng ký {g2}",
        f"DoanhSo_{g1}": f"Doanh số tích lũy {g1}",
        f"DoanhSo_{g2}": f"Doanh số tích lũy {g2}",
        f"Nguong_{g2}": "Ngưỡng tối thiểu",
        "KetQua": "Kết quả", "GhiChu": "Ghi chú",
    }
    if file_t0:
        rename_cols[f"SoSuat_{g0}"] = f"Số suất đăng ký {g0}"
        rename_cols[f"DoanhSo_{g0}"] = f"Doanh số tích lũy {g0}"

    df_out = df_final[cols_output].copy().rename(columns=rename_cols)
    df_removed_out = df_removed[cols_output].copy().rename(columns=rename_cols)

    # --- GSBH notes: chỉ để lại "Thiếu: xxx" cho Không đạt ---
    def fmt_v(v): return f"{int(round(float(v))):,}".replace(",", ".")
    if "Ngưỡng tối thiểu" in df_out.columns:
        doanh_so_cols = sorted([c for c in df_out.columns if c.startswith("Doanh số tích lũy ")])
        if doanh_so_cols:
            col_ds_t2 = doanh_so_cols[-1]
            mask_nd = df_out["Kết quả"].eq("Không đạt")
            remain = (df_out.loc[mask_nd, "Ngưỡng tối thiểu"].astype(float) - df_out.loc[mask_nd, col_ds_t2].astype(float)).clip(lower=0)
            df_out.loc[mask_nd, "Ghi chú"] = remain.map(lambda v: f"Thiếu: {fmt_v(v)}")

    df_out = df_out.sort_values(by=["Mã NPP", "Mã NVBH", "Tên khách hàng"])
    df_removed_out = df_removed_out.sort_values(by=["Mã NPP", "Mã NVBH", "Tên khách hàng"])
    return df_out, df_removed_out

# =============== Streamlit UI ===============
st.set_page_config(page_title="Cholimex Display Checker", layout="wide")
st.title("Cholimex Foods Display Checker (Web)")
st.caption("Upload file Excel đầu vào, chọn bộ lọc và tải file kết quả.")

with st.expander("⚙️ Thiết lập"):
    colA, colB, colC = st.columns(3)
    with colA:
        mode = st.selectbox("Chế độ xuất", ["MKT", "GSBH"], index=0)
    with colB:
        regions = list(REGION_MAP.keys())
        sel_regions = st.multiselect("Miền xuất báo cáo", regions, default=[])
    with colC:
        st.markdown("**Bộ lọc Kết quả**")
        kq_all = st.checkbox("Tất cả", value=False)
        kq_dat = st.checkbox("Đạt", value=False)
        kq_khongdat = st.checkbox("Không đạt", value=False)
        kq_khongxet = st.checkbox("Không xét", value=False)

st.markdown("**1) Upload các file Excel (có cột _Mức đăng ký_ và _Giai đoạn_)**")
uploads = st.file_uploader("Chọn nhiều file .xlsx/.xls", type=["xlsx", "xls"], accept_multiple_files=True)

# --- Nhận diện CT & tháng trong từng file ---
file_entries = []
err_msgs = []

if uploads:
    for f in uploads:
        data = f.read()
        bio = io.BytesIO(data)
        try:
            # Đọc nhanh để bắt CT & Giai đoạn
            df_preview = pd.read_excel(io.BytesIO(data), header=1)
            ct = detect_ct_from_content(df_preview)
            if not ct:
                ct = detect_ct_from_filename(f.name)  # fallback
            if not ct:
                err_msgs.append(f"- Không nhận diện được CT từ: {f.name} (thiếu 'Mức đăng ký' hoặc tên file không chứa mã CT).")
                continue

            if "Giai đoạn" not in df_preview.columns:
                err_msgs.append(f"- File {f.name} thiếu cột 'Giai đoạn'.")
                continue
            giai_val = str(df_preview["Giai đoạn"].dropna().iloc[0]).strip()
            dt = parse_giai_to_dt(giai_val)

            file_entries.append({
                "name": f.name,
                "bytes": data,      # lưu bytes để đọc lại nhiều lần
                "ct": ct,
                "giai": giai_val,
                "dt": dt
            })
        except Exception as e:
            err_msgs.append(f"- Lỗi đọc {f.name}: {e}")

# --- Scan 'Thứ bán hàng' từ tháng mới nhất từng CT ---
route_values = set()
if file_entries:
    by_ct = defaultdict(list)
    for ent in file_entries:
        by_ct[ent["ct"]].append(ent)
    for ct, items in by_ct.items():
        last = sorted(items, key=lambda x: x["dt"])[-1]
        try:
            df_last = pd.read_excel(io.BytesIO(last["bytes"]), header=1)
            col = None
            for cand in ["Thứ bán hàng", "ThuBanHang", "Tuyến", "Tuyen", "Route", "Tuyến bán hàng", "Tuyến BH", "Mã tuyến", "T/BH", "TBH"]:
                if cand in df_last.columns:
                    col = cand
                    break
            if col is not None:
                vals = df_last[col].dropna().astype(str).map(lambda s: s.strip())
                route_values.update([v for v in vals if v])
        except Exception as e:
            err_msgs.append(f"- Lỗi quét 'Thứ bán hàng' CT {ct}: {e}")

st.markdown("**2) Lọc theo ‘Thứ bán hàng’ (tuỳ chọn)**")
sel_routes = st.multiselect("Chọn nhiều", sorted(route_values, key=lambda s: s.upper()))

if err_msgs:
    st.warning("⚠️ Một số tệp gặp vấn đề:\n" + "\n".join(err_msgs))

run = st.button("▶︎ Xuất báo cáo")

if run:
    if not file_entries:
        st.error("Vui lòng upload ít nhất 2 tệp hợp lệ."); st.stop()
    if not sel_regions:
        st.error("Vui lòng chọn ít nhất 1 miền."); st.stop()

    # Gom theo CT và sắp theo tháng
    by_ct = defaultdict(list)
    for ent in file_entries:
        by_ct[ent["ct"]].append(ent)

    # Kết quả theo từng miền (tạo file Excel trong memory)
    outputs_per_region = {}     # region -> BytesIO
    outputs_xoa_per_region = {} # chỉ MKT

    # Chọn bộ lọc kết quả
    if kq_all:
        selected_kq = None
    else:
        selected_kq = set()
        if kq_dat: selected_kq.add("Đạt")
        if kq_khongdat: selected_kq.add("Không đạt")
        if kq_khongxet: selected_kq.add("Không xét")
        if not selected_kq:
            selected_kq = None

    for region in sel_regions:
        bao_cao_data, bao_cao_huy = [], []

        # Writer chính
        bio_main = io.BytesIO()
        writer_main = pd.ExcelWriter(bio_main, engine="openpyxl")
        # Writer XÓA (MKT)
        writer_xoa = None
        bio_xoa = None
        if mode != "GSBH":
            bio_xoa = io.BytesIO()
            writer_xoa = pd.ExcelWriter(bio_xoa, engine="openpyxl")

        idx = 0
        for ct, items in by_ct.items():
            items_sorted = sorted(items, key=lambda x: x["dt"])
            # cần >=2 tháng
            if len(items_sorted) < 2:
                continue
            # lấy T1, T2 (2 tháng cuối) và optional T0
            t2 = items_sorted[-1]
            t1 = items_sorted[-2]
            t0 = items_sorted[0] if len(items_sorted) >= 3 else None

            # Process
            try:
                df_out, df_removed = xu_ly_chuong_trinh(
                    file_truoc=io.BytesIO(t1["bytes"]),
                    file_sau=io.BytesIO(t2["bytes"]),
                    muc_toi_thieu=MUC_TOI_THIEU,
                    program_names=PROGRAM_NAMES,
                    xbm_map=XBM_MAP,
                    file_t0=(io.BytesIO(t0["bytes"]) if t0 else None),
                    filter_ketqua=selected_kq,
                    filter_tuyen_tokens=(sel_routes if sel_routes else None),
                )
            except Exception as e:
                st.error(f"Lỗi xử lý CT {ct}: {e}")
                continue

            # Lọc theo miền
            if REGION_MAP.get(region) != "ALL":
                df_out = df_out[df_out["Miền"].isin(REGION_MAP[region])]
                df_removed = df_removed[df_removed["Miền"].isin(REGION_MAP[region])]

            # GSBH: rút gọn cột + giữ định dạng
            if mode == "GSBH":
                keep_cols = ["Mức đăng ký", "Tên NPP", "Mã NVBH", "Tên NVBH",
                             "Mã khách hàng", "Tên khách hàng", "Thứ bán hàng"]
                so_suat_cols = sorted([c for c in df_out.columns if c.startswith("Số suất đăng ký ")])
                doanh_so_cols = sorted([c for c in df_out.columns if c.startswith("Doanh số tích lũy ")])
                if len(so_suat_cols) >= 2: keep_cols += [so_suat_cols[-2], so_suat_cols[-1]]
                elif len(so_suat_cols) == 1: keep_cols += [so_suat_cols[-1]]
                if len(doanh_so_cols) >= 2: keep_cols += [doanh_so_cols[-2], doanh_so_cols[-1]]
                elif len(doanh_so_cols) == 1: keep_cols += [doanh_so_cols[-1]]
                keep_cols += ["Ngưỡng tối thiểu", "Kết quả", "Ghi chú"]
                keep_cols = [c for c in keep_cols if c in df_out.columns]
                df_out = df_out[keep_cols]

            # Ghi sheet
            df_out.to_excel(writer_main, sheet_name=ct, index=False)
            style_excel(writer_main, ct)
            if mode != "GSBH" and writer_xoa is not None:
                df_removed.to_excel(writer_xoa, sheet_name=ct, index=False)
                style_excel(writer_xoa, ct)

            # Tổng hợp
            try:
                idx += 1
                tong_suat = df_out.filter(like="Số suất đăng ký").iloc[:, -1].sum()
                ko_dat = df_out.loc[df_out["Kết quả"] == "Không đạt", :].filter(like="Số suất đăng ký").iloc[:, -1].sum()
                tile = f"{(ko_dat / tong_suat):.1%}" if tong_suat > 0 else "0%"
                bao_cao_data.append([idx, PROGRAM_NAMES.get(ct, ct), MUC_TOI_THIEU.get(ct, 0), int(tong_suat), int(ko_dat), tile])
                if mode != "GSBH":
                    so_huy = df_out.loc[df_out["Kết quả"] == "Không đạt", :].filter(like="Số suất đăng ký").iloc[:, -1].sum()
                    bao_cao_huy.append([idx, PROGRAM_NAMES.get(ct, ct), int(so_huy)])
            except Exception as e:
                st.warning(f"⚠️ Lỗi thống kê CT {ct}: {e}")

        # Sheet tổng hợp
        if bao_cao_data:
            try: tao_bao_cao_tonghop(writer_main, bao_cao_data)
            except Exception as e: st.warning(f"⚠️ Lỗi tạo BaoCao_TongHop: {e}")
        if (mode != "GSBH") and bao_cao_huy:
            try: tao_bao_cao_huy(writer_main, bao_cao_huy)
            except Exception as e: st.warning(f"⚠️ Lỗi tạo BaoCao_Huy: {e}")

        writer_main.close()
        outputs_per_region[region] = bio_main.getvalue()
        if mode != "GSBH" and writer_xoa is not None:
            writer_xoa.close()
            outputs_xoa_per_region[region] = bio_xoa.getvalue()

    st.success("✅ Đã xử lý xong.")
    # Tải từng file hoặc ZIP
    cols = st.columns(2)
    with cols[0]:
        st.markdown("**Tải từng miền**")
        for region, data in outputs_per_region.items():
            label = f"⬇️ Tải {region} ({'GSBH' if mode=='GSBH' else 'MKT'})"
            st.download_button(label, data=data, file_name=f"TongHop_{region}{'_GSBH' if mode=='GSBH' else ''}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if mode != "GSBH":
            st.markdown("**Tải file XÓA (MKT)**")
            for region, data in outputs_xoa_per_region.items():
                st.download_button(f"⬇️ Tải Xóa {region}", data=data, file_name=f"TongHop_Xoa_{region}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with cols[1]:
        if outputs_per_region:
            st.markdown("**Tải tất cả dưới dạng ZIP**")
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for region, data in outputs_per_region.items():
                    zf.writestr(f"TongHop_{region}{'_GSBH' if mode=='GSBH' else ''}.xlsx", data)
                if mode != "GSBH":
                    for region, data in outputs_xoa_per_region.items():
                        zf.writestr(f"TongHop_Xoa_{region}.xlsx", data)
            st.download_button("⬇️ Tải tất cả (.zip)", data=zip_buf.getvalue(), file_name="KetQua_TongHop.zip", mime="application/zip")
